#!/usr/bin/env python3
"""
创建测试Excel文件用于导入测试
"""

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os

def create_test_excel():
    """创建测试Excel文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "SKC测试数据"
    
    # 测试数据
    test_data = {
        "产品A": [
            ("SKC001", "核价通过"),
            ("SKC002", "拉过库存"),
            ("SKC003", "已下架"),
        ],
        "产品B": [
            ("SKC004", "价格待定"),
            ("SKC005", "核价通过"),
            ("SKC006", "减少库存为0"),
        ],
        "产品C": [
            ("SKC007", "改过体积"),
            ("SKC008", "价格错误"),
            ("SKC009", "核价通过"),
        ]
    }
    
    col = 1
    for product_name, skcs in test_data.items():
        # 写产品名（合并单元格）
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+1)
        ws.cell(row=1, column=col, value=product_name)
        
        # 写表头
        ws.cell(row=3, column=col, value="SKC")
        ws.cell(row=3, column=col+1, value="状态")
        
        # 写SKC数据
        for i, (skc_code, status) in enumerate(skcs, start=4):
            ws.cell(row=i, column=col, value=skc_code)
            ws.cell(row=i, column=col+1, value=status)
        
        col += 2
    
    # 保存文件
    filename = "test_skc_data.xlsx"
    wb.save(filename)
    print(f"✅ 测试Excel文件已创建: {filename}")
    print(f"📁 文件位置: {os.path.abspath(filename)}")
    
    return filename

if __name__ == "__main__":
    create_test_excel()
    print("\n📋 使用说明:")
    print("1. 在Web界面中登录系统")
    print("2. 创建或选择一个项目")
    print("3. 点击'导入Excel'按钮")
    print("4. 选择刚创建的 test_skc_data.xlsx 文件")
    print("5. 等待导入完成")