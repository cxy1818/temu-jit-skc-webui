from flask import Blueprint, request, jsonify, current_app, send_file
from flask_login import login_required, current_user
from models import db, Project, Product, SKC, ProductImage, ExcelExport, STATUS_OPTIONS
from werkzeug.utils import secure_filename
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import os
import time
import uuid
from datetime import datetime
from PIL import Image
import io

api_bp = Blueprint('api', __name__, url_prefix='/api')

def allowed_file(filename, allowed_extensions):
    """检查文件扩展名是否允许"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in allowed_extensions

def save_uploaded_file(file, upload_folder, allowed_extensions):
    """保存上传的文件"""
    if file and allowed_file(file.filename, allowed_extensions):
        # 生成唯一文件名
        filename = secure_filename(file.filename)
        name, ext = os.path.splitext(filename)
        unique_filename = f"{name}_{uuid.uuid4().hex[:8]}{ext}"
        
        # 确保上传目录存在
        os.makedirs(upload_folder, exist_ok=True)
        
        file_path = os.path.join(upload_folder, unique_filename)
        file.save(file_path)
        
        return unique_filename, file_path
    return None, None

# ========== 项目管理 API ==========

@api_bp.route('/projects', methods=['GET'])
@login_required
def get_projects():
    """获取用户的所有项目"""
    page = request.args.get('page', 1, type=int)
    per_page = min(request.args.get('per_page', 20, type=int), 100)
    
    projects = Project.query.filter_by(
        user_id=current_user.id, 
        is_active=True
    ).order_by(Project.updated_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    return jsonify({
        'success': True,
        'projects': [{
            'id': p.id,
            'name': p.name,
            'description': p.description,
            'created_at': p.created_at.isoformat(),
            'updated_at': p.updated_at.isoformat(),
            'product_count': p.products.count()
        } for p in projects.items],
        'pagination': {
            'page': projects.page,
            'pages': projects.pages,
            'per_page': projects.per_page,
            'total': projects.total
        }
    })

@api_bp.route('/projects', methods=['POST'])
@login_required
def create_project():
    """创建新项目"""
    data = request.get_json()
    name = data.get('name', '').strip()
    description = data.get('description', '').strip()
    
    if not name:
        return jsonify({'success': False, 'message': '项目名称不能为空'}), 400
    
    # 检查项目名是否已存在
    existing = Project.query.filter_by(
        user_id=current_user.id, 
        name=name, 
        is_active=True
    ).first()
    
    if existing:
        return jsonify({'success': False, 'message': '项目名称已存在'}), 400
    
    try:
        project = Project(
            name=name,
            description=description,
            user_id=current_user.id
        )
        db.session.add(project)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': '项目创建成功',
            'project': {
                'id': project.id,
                'name': project.name,
                'description': project.description,
                'created_at': project.created_at.isoformat()
            }
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': '创建项目失败'}), 500

@api_bp.route('/projects/<int:project_id>', methods=['PUT'])
@login_required
def update_project(project_id):
    """更新项目信息"""
    project = Project.query.filter_by(
        id=project_id, 
        user_id=current_user.id, 
        is_active=True
    ).first()
    
    if not project:
        return jsonify({'success': False, 'message': '项目不存在'}), 404
    
    data = request.get_json()
    name = data.get('name', '').strip()
    description = data.get('description', '').strip()
    
    if not name:
        return jsonify({'success': False, 'message': '项目名称不能为空'}), 400
    
    # 检查项目名是否已存在（排除当前项目）
    existing = Project.query.filter(
        Project.user_id == current_user.id,
        Project.name == name,
        Project.is_active == True,
        Project.id != project_id
    ).first()
    
    if existing:
        return jsonify({'success': False, 'message': '项目名称已存在'}), 400
    
    try:
        project.name = name
        project.description = description
        project.updated_at = datetime.utcnow()
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': '项目更新成功',
            'project': {
                'id': project.id,
                'name': project.name,
                'description': project.description,
                'updated_at': project.updated_at.isoformat()
            }
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': '更新项目失败'}), 500

@api_bp.route('/projects/<int:project_id>', methods=['DELETE'])
@login_required
def delete_project(project_id):
    """删除项目（软删除）"""
    project = Project.query.filter_by(
        id=project_id, 
        user_id=current_user.id, 
        is_active=True
    ).first()
    
    if not project:
        return jsonify({'success': False, 'message': '项目不存在'}), 404
    
    try:
        project.is_active = False
        project.updated_at = datetime.utcnow()
        db.session.commit()
        
        return jsonify({'success': True, 'message': '项目删除成功'})
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': '删除项目失败'}), 500

# ========== 产品管理 API ==========

@api_bp.route('/projects/<int:project_id>/products', methods=['GET'])
@login_required
def get_products(project_id):
    """获取项目的所有产品"""
    project = Project.query.filter_by(
        id=project_id, 
        user_id=current_user.id, 
        is_active=True
    ).first()
    
    if not project:
        return jsonify({'success': False, 'message': '项目不存在'}), 404
    
    page = request.args.get('page', 1, type=int)
    per_page = min(request.args.get('per_page', 20, type=int), 100)
    
    products = Product.query.filter_by(
        project_id=project_id
    ).order_by(Product.updated_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    return jsonify({
        'success': True,
        'products': [{
            'id': p.id,
            'name': p.name,
            'created_at': p.created_at.isoformat(),
            'updated_at': p.updated_at.isoformat(),
            'skc_count': p.skcs.count(),
            'image_count': p.images.count()
        } for p in products.items],
        'pagination': {
            'page': products.page,
            'pages': products.pages,
            'per_page': products.per_page,
            'total': products.total
        }
    })

@api_bp.route('/projects/<int:project_id>/products', methods=['POST'])
@login_required
def create_product(project_id):
    """创建新产品"""
    project = Project.query.filter_by(
        id=project_id, 
        user_id=current_user.id, 
        is_active=True
    ).first()
    
    if not project:
        return jsonify({'success': False, 'message': '项目不存在'}), 404
    
    data = request.get_json()
    name = data.get('name', '').strip()
    
    if not name:
        return jsonify({'success': False, 'message': '产品名称不能为空'}), 400
    
    # 检查产品名是否已存在
    existing = Product.query.filter_by(
        project_id=project_id, 
        name=name
    ).first()
    
    if existing:
        return jsonify({'success': False, 'message': '产品名称已存在'}), 400
    
    try:
        product = Product(
            name=name,
            project_id=project_id
        )
        db.session.add(product)
        
        # 更新项目的更新时间
        project.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': '产品创建成功',
            'product': {
                'id': product.id,
                'name': product.name,
                'created_at': product.created_at.isoformat()
            }
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': '创建产品失败'}), 500

# ========== SKC管理 API ==========

@api_bp.route('/products/<int:product_id>/skcs', methods=['GET'])
@login_required
def get_skcs(product_id):
    """获取产品的所有SKC"""
    product = Product.query.join(Project).filter(
        Product.id == product_id,
        Project.user_id == current_user.id,
        Project.is_active == True
    ).first()
    
    if not product:
        return jsonify({'success': False, 'message': '产品不存在'}), 404
    
    page = request.args.get('page', 1, type=int)
    per_page = min(request.args.get('per_page', 50, type=int), 200)
    status_filter = request.args.get('status')
    
    query = SKC.query.filter_by(product_id=product_id)
    
    if status_filter and status_filter in STATUS_OPTIONS:
        query = query.filter_by(status=status_filter)
    
    skcs = query.order_by(
        db.case(
            *[(SKC.status == status, idx) for idx, status in enumerate(STATUS_OPTIONS)],
            else_=len(STATUS_OPTIONS)
        ),
        SKC.code
    ).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    return jsonify({
        'success': True,
        'skcs': [{
            'id': s.id,
            'code': s.code,
            'status': s.status,
            'created_at': s.created_at.isoformat(),
            'updated_at': s.updated_at.isoformat()
        } for s in skcs.items],
        'pagination': {
            'page': skcs.page,
            'pages': skcs.pages,
            'per_page': skcs.per_page,
            'total': skcs.total
        },
        'status_options': STATUS_OPTIONS
    })

@api_bp.route('/products/<int:product_id>/skcs', methods=['POST'])
@login_required
def add_skcs(product_id):
    """批量添加SKC"""
    product = Product.query.join(Project).filter(
        Product.id == product_id,
        Project.user_id == current_user.id,
        Project.is_active == True
    ).first()
    
    if not product:
        return jsonify({'success': False, 'message': '产品不存在'}), 404
    
    data = request.get_json()
    skc_codes = data.get('skc_codes', [])
    status = data.get('status', '核价通过')
    
    if not skc_codes:
        return jsonify({'success': False, 'message': 'SKC代码不能为空'}), 400
    
    if status not in STATUS_OPTIONS:
        return jsonify({'success': False, 'message': '状态选项无效'}), 400
    
    try:
        added_count = 0
        duplicate_codes = []
        
        for code in skc_codes:
            code = str(code).strip()
            if not code:
                continue
            
            # 检查SKC是否已存在（全局唯一）
            existing = SKC.query.filter_by(code=code).first()
            if existing:
                duplicate_codes.append(code)
                continue
            
            skc = SKC(
                code=code,
                status=status,
                product_id=product_id
            )
            db.session.add(skc)
            added_count += 1
        
        # 更新产品和项目的更新时间
        product.updated_at = datetime.utcnow()
        product.project.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        message = f'成功添加 {added_count} 个SKC'
        if duplicate_codes:
            message += f'，跳过重复的SKC: {", ".join(duplicate_codes[:5])}'
            if len(duplicate_codes) > 5:
                message += f' 等{len(duplicate_codes)}个'
        
        return jsonify({
            'success': True,
            'message': message,
            'added_count': added_count,
            'duplicate_count': len(duplicate_codes)
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': '添加SKC失败'}), 500

@api_bp.route('/skcs/batch_update', methods=['PUT'])
@login_required
def batch_update_skcs():
    """批量更新SKC状态"""
    data = request.get_json()
    skc_codes = data.get('skc_codes', [])
    new_status = data.get('status')
    
    if not skc_codes:
        return jsonify({'success': False, 'message': 'SKC代码不能为空'}), 400
    
    if new_status not in STATUS_OPTIONS:
        return jsonify({'success': False, 'message': '状态选项无效'}), 400
    
    try:
        # 查找用户有权限的SKC
        skcs = SKC.query.join(Product).join(Project).filter(
            SKC.code.in_(skc_codes),
            Project.user_id == current_user.id,
            Project.is_active == True
        ).all()
        
        if not skcs:
            return jsonify({'success': False, 'message': '未找到可更新的SKC'}), 404
        
        updated_count = 0
        for skc in skcs:
            skc.status = new_status
            skc.updated_at = datetime.utcnow()
            skc.product.updated_at = datetime.utcnow()
            skc.product.project.updated_at = datetime.utcnow()
            updated_count += 1
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'成功更新 {updated_count} 个SKC状态为「{new_status}」',
            'updated_count': updated_count
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': '更新SKC失败'}), 500

@api_bp.route('/skcs/batch_delete', methods=['DELETE'])
@login_required
def batch_delete_skcs():
    """批量删除SKC"""
    data = request.get_json()
    skc_codes = data.get('skc_codes', [])
    
    if not skc_codes:
        return jsonify({'success': False, 'message': 'SKC代码不能为空'}), 400
    
    try:
        # 查找用户有权限的SKC
        skcs = SKC.query.join(Product).join(Project).filter(
            SKC.code.in_(skc_codes),
            Project.user_id == current_user.id,
            Project.is_active == True
        ).all()
        
        if not skcs:
            return jsonify({'success': False, 'message': '未找到可删除的SKC'}), 404
        
        deleted_count = len(skcs)
        
        for skc in skcs:
            skc.product.updated_at = datetime.utcnow()
            skc.product.project.updated_at = datetime.utcnow()
            db.session.delete(skc)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'成功删除 {deleted_count} 个SKC',
            'deleted_count': deleted_count
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': '删除SKC失败'}), 500

# ========== 图片管理 API ==========

# ========== 图片管理 API ==========

@api_bp.route('/products/<int:product_id>/images', methods=['GET'])
@login_required
def get_product_images(product_id):
    """获取产品的所有图片"""
    product = Product.query.join(Project).filter(
        Product.id == product_id,
        Project.user_id == current_user.id,
        Project.is_active == True
    ).first()
    
    if not product:
        return jsonify({'success': False, 'message': '产品不存在'}), 404
    
    images = ProductImage.query.filter_by(product_id=product_id).order_by(
        ProductImage.is_primary.desc(),
        ProductImage.uploaded_at.desc()
    ).all()
    
    return jsonify({
        'success': True,
        'images': [{
            'id': img.id,
            'filename': img.filename,
            'original_filename': img.original_filename,
            'file_path': img.file_path,
            'file_size': img.file_size,
            'mime_type': img.mime_type,
            'is_primary': img.is_primary,
            'uploaded_at': img.uploaded_at.isoformat()
        } for img in images]
    })

@api_bp.route('/products/<int:product_id>/images', methods=['POST'])
@login_required
def upload_product_image(product_id):
    """上传产品图片"""
    product = Product.query.join(Project).filter(
        Product.id == product_id,
        Project.user_id == current_user.id,
        Project.is_active == True
    ).first()
    
    if not product:
        return jsonify({'success': False, 'message': '产品不存在'}), 404
    
    if 'image' not in request.files:
        return jsonify({'success': False, 'message': '未选择图片文件'}), 400
    
    file = request.files['image']
    if file.filename == '':
        return jsonify({'success': False, 'message': '未选择图片文件'}), 400
    
    # 保存图片文件
    upload_folder = os.path.join(current_app.config['UPLOAD_FOLDER'], 'images')
    filename, file_path = save_uploaded_file(
        file, upload_folder, {'png', 'jpg', 'jpeg', 'gif'}
    )
    
    if not filename:
        return jsonify({'success': False, 'message': '图片格式不支持'}), 400
    
    try:
        # 获取文件信息
        file_size = os.path.getsize(file_path)
        
        # 创建图片记录
        image = ProductImage(
            filename=filename,
            original_filename=file.filename,
            file_path=file_path,
            file_size=file_size,
            mime_type=file.mimetype,
            product_id=product_id,
            is_primary=product.images.count() == 0  # 第一张图片设为主图
        )
        
        db.session.add(image)
        
        # 更新产品和项目的更新时间
        product.updated_at = datetime.utcnow()
        product.project.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': '图片上传成功',
            'image': {
                'id': image.id,
                'filename': image.filename,
                'original_filename': image.original_filename,
                'is_primary': image.is_primary,
                'uploaded_at': image.uploaded_at.isoformat()
            }
        })
    
    except Exception as e:
        db.session.rollback()
        # 删除已上传的文件
        if os.path.exists(file_path):
            os.remove(file_path)
        return jsonify({'success': False, 'message': '保存图片失败'}), 500

@api_bp.route('/images/<int:image_id>/primary', methods=['PUT'])
@login_required
def set_primary_image(image_id):
    """设置图片为主图"""
    image = ProductImage.query.join(Product).join(Project).filter(
        ProductImage.id == image_id,
        Project.user_id == current_user.id,
        Project.is_active == True
    ).first()
    
    if not image:
        return jsonify({'success': False, 'message': '图片不存在'}), 404
    
    try:
        # 取消该产品的其他主图
        ProductImage.query.filter_by(
            product_id=image.product_id,
            is_primary=True
        ).update({'is_primary': False})
        
        # 设置当前图片为主图
        image.is_primary = True
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': '主图设置成功'
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': '设置主图失败'}), 500

@api_bp.route('/images/<int:image_id>', methods=['DELETE'])
@login_required
def delete_image(image_id):
    """删除图片"""
    image = ProductImage.query.join(Product).join(Project).filter(
        ProductImage.id == image_id,
        Project.user_id == current_user.id,
        Project.is_active == True
    ).first()
    
    if not image:
        return jsonify({'success': False, 'message': '图片不存在'}), 404
    
    try:
        # 删除文件
        if os.path.exists(image.file_path):
            os.remove(image.file_path)
        
        # 删除数据库记录
        db.session.delete(image)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': '图片删除成功'
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': '删除图片失败'}), 500

@api_bp.route('/stats/user', methods=['GET'])
@login_required
def get_user_stats():
    """获取用户的统计数据"""
    try:
        # 获取用户的项目数
        project_count = Project.query.filter_by(
            user_id=current_user.id,
            is_active=True
        ).count()
        
        # 获取用户的产品数
        product_count = Product.query.join(Project).filter(
            Project.user_id == current_user.id,
            Project.is_active == True
        ).count()
        
        # 获取用户的SKC数
        skc_count = SKC.query.join(Product).join(Project).filter(
            Project.user_id == current_user.id,
            Project.is_active == True
        ).count()
        
        # 获取用户的图片数
        image_count = ProductImage.query.join(Product).join(Project).filter(
            Project.user_id == current_user.id,
            Project.is_active == True
        ).count()
        
        return jsonify({
            'success': True,
            'stats': {
                'project_count': project_count,
                'product_count': product_count,
                'skc_count': skc_count,
                'image_count': image_count
            }
        })
    
    except Exception as e:
        return jsonify({'success': False, 'message': '获取统计数据失败'}), 500

# ========== Excel导入导出 API ==========

@api_bp.route('/projects/<int:project_id>/import', methods=['POST'])
@login_required
def import_excel_data(project_id):
    """导入Excel数据"""
    project = Project.query.filter_by(
        id=project_id, 
        user_id=current_user.id, 
        is_active=True
    ).first()
    
    if not project:
        return jsonify({'success': False, 'message': '项目不存在'}), 404
    
    if 'excel' not in request.files:
        return jsonify({'success': False, 'message': '未选择Excel文件'}), 400
    
    file = request.files['excel']
    if file.filename == '':
        return jsonify({'success': False, 'message': '未选择Excel文件'}), 400
    
    # 保存Excel文件
    upload_folder = os.path.join(current_app.config['UPLOAD_FOLDER'], 'temp')
    filename, file_path = save_uploaded_file(
        file, upload_folder, {'xlsx', 'xlsm'}
    )
    
    if not filename:
        return jsonify({'success': False, 'message': 'Excel格式不支持'}), 400
    
    try:
        # 读取Excel文件
        wb = load_workbook(file_path, data_only=True)
        imported_count = 0
        skipped_count = 0
        
        for ws in wb.worksheets:
            max_col = ws.max_column
            
            # 按列处理数据（每两列为一个产品）
            for col in range(1, max_col + 1, 2):
                product_name_cell = ws.cell(row=1, column=col)
                if not product_name_cell.value:
                    continue
                
                product_name = str(product_name_cell.value).strip()
                
                # 创建或获取产品
                product = Product.query.filter_by(
                    project_id=project_id,
                    name=product_name
                ).first()
                
                if not product:
                    product = Product(
                        name=product_name,
                        project_id=project_id
                    )
                    db.session.add(product)
                    db.session.flush()  # 获取ID
                
                # 读取SKC数据（从第4行开始）
                max_row = ws.max_row
                for row in range(4, max_row + 1):
                    skc_cell = ws.cell(row=row, column=col)
                    status_cell = ws.cell(row=row, column=col + 1)
                    
                    if not skc_cell.value or not status_cell.value:
                        continue
                    
                    skc_code = str(skc_cell.value).strip()
                    status = str(status_cell.value).strip()
                    
                    # 检查SKC是否已存在
                    existing_skc = SKC.query.filter_by(code=skc_code).first()
                    if existing_skc:
                        skipped_count += 1
                        continue
                    
                    # 验证状态
                    if status not in STATUS_OPTIONS:
                        status = '核价通过'  # 默认状态
                    
                    # 创建SKC
                    skc = SKC(
                        code=skc_code,
                        status=status,
                        product_id=product.id
                    )
                    db.session.add(skc)
                    imported_count += 1
        
        # 更新项目时间
        project.updated_at = datetime.utcnow()
        db.session.commit()
        
        # 删除临时文件
        if os.path.exists(file_path):
            os.remove(file_path)
        
        message = f'成功导入 {imported_count} 条记录'
        if skipped_count > 0:
            message += f'，跳过重复记录 {skipped_count} 条'
        
        return jsonify({
            'success': True,
            'message': message,
            'imported_count': imported_count,
            'skipped_count': skipped_count
        })
    
    except Exception as e:
        db.session.rollback()
        # 删除临时文件
        if os.path.exists(file_path):
            os.remove(file_path)
        return jsonify({'success': False, 'message': f'导入失败: {str(e)}'}), 500

@api_bp.route('/projects/<int:project_id>/export', methods=['POST'])
@login_required
def export_project_excel(project_id):
    """导出项目为Excel文件"""
    project = Project.query.filter_by(
        id=project_id, 
        user_id=current_user.id, 
        is_active=True
    ).first()
    
    if not project:
        return jsonify({'success': False, 'message': '项目不存在'}), 404
    
    try:
        # 创建Excel文件
        wb = Workbook()
        ws = wb.active
        ws.title = project.name
        
        # 获取项目的所有产品和SKC
        products = Product.query.filter_by(project_id=project_id).all()
        
        col = 1
        for product in products:
            # 写产品名
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+1)
            ws.cell(row=1, column=col, value=product.name)
            
            # 添加图片（如果有主图）
            primary_image = ProductImage.query.filter_by(
                product_id=product.id, 
                is_primary=True
            ).first()
            
            if primary_image and os.path.exists(primary_image.file_path):
                try:
                    img = XLImage(primary_image.file_path)
                    img.width = 100
                    img.height = 100
                    ws.add_image(img, f"{get_column_letter(col)}2")
                    ws.row_dimensions[2].height = 80
                    ws.column_dimensions[get_column_letter(col)].width = 15
                except Exception:
                    pass
            
            # 写表头
            ws.cell(row=3, column=col, value="SKC")
            ws.cell(row=3, column=col+1, value="状态")
            
            # 写SKC数据
            skcs = SKC.query.filter_by(product_id=product.id).order_by(
                db.case(
                    *[(SKC.status == status, idx) for idx, status in enumerate(STATUS_OPTIONS)],
                    else_=len(STATUS_OPTIONS)
                )
            ).all()
            
            row = 4
            for skc in skcs:
                ws.cell(row=row, column=col, value=skc.code)
                ws.cell(row=row, column=col+1, value=skc.status)
                row += 1
            
            col += 2
        
        # 保存文件
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        filename = f"{project.name}_{timestamp}.xlsx"
        safe_filename = secure_filename(filename)
        
        export_folder = os.path.join(current_app.config['UPLOAD_FOLDER'], 'exports')
        os.makedirs(export_folder, exist_ok=True)
        
        file_path = os.path.join(export_folder, safe_filename)
        wb.save(file_path)
        
        # 记录导出历史
        file_size = os.path.getsize(file_path)
        export_record = ExcelExport(
            filename=safe_filename,
            file_path=file_path,
            project_id=project_id,
            user_id=current_user.id,
            file_size=file_size
        )
        
        db.session.add(export_record)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Excel导出成功',
            'export': {
                'id': export_record.id,
                'filename': export_record.filename,
                'created_at': export_record.created_at.isoformat(),
                'file_size': export_record.file_size
            }
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'导出失败: {str(e)}'}), 500

@api_bp.route('/exports/<int:export_id>/download')
@login_required
def download_export(export_id):
    """下载导出的Excel文件"""
    export = ExcelExport.query.filter_by(
        id=export_id,
        user_id=current_user.id
    ).first()
    
    if not export:
        return jsonify({'success': False, 'message': '文件不存在'}), 404
    
    if not os.path.exists(export.file_path):
        return jsonify({'success': False, 'message': '文件已被删除'}), 404
    
    return send_file(
        export.file_path,
        as_attachment=True,
        download_name=export.filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )